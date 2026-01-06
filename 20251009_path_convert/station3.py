# -*- coding: utf-8 -*-
"""
Created on Mon Mar 11 10:40:38 2024

@author: thinkpad
"""

import ezdxf
import openpyxl
import re


def filter_strings(keywords, content):
    pattern = keywords
    result = []
    content_asm = [content]
    for string in content_asm:
        if re.search(pattern, string):
#            string = list(string)
#            for i in range(len(string)):
#                if(string[i] == 'R' or string[i] == 'L'):
#                    if(string[i+1] == '-' and string[i+2] != ' '):
#                        result = string[i+2]
#                        for j in range(i+2, len(string), 1)
            string = re.sub(pattern, '', string)
            string = re.split(r'[+]', string)     
           
            result.extend(string)
           
    return result

def station(Load, Save, x0, x1, y0, y1, lay, k_DK):
    #打开DWG文件
    dwg_file = Load
    doc = ezdxf.readfile(dwg_file)
    msp = doc.modelspace()
    #设施数据区域
    x_min = x0
    x_max = x1
    y_min = y0 
    y_max = y1
    layer = lay
#    layer_text = lay_t
    #创建一个新的Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "线路设施"
    #创建表头
    ws['A1'] = "设施名称"
    ws['B1'] = "里程"
    station = []  #站名数组
    position = []  #公里标数组
    vline_x = []  #垂直线x坐标数组
    station_count = 0  #站点计数器
    v_count = 0  #内循环计数器
    #筛选直线    
    lines = msp.query('LINE[layer=="%s"]' % layer)
    for line in lines:
        x_start = line.dxf.start[0]
        y_start = line.dxf.start[1]
        x_end = line.dxf.end[0]
        y_end = line.dxf.end[1]
        if(x_min <= x_start <= x_max and y_min < y_start < y_max and x_min <= 
           x_end <= x_max and y_min < y_end < y_max):
            if(round(x_start, 8) == round(x_end, 8)):  #筛选垂直线
                vline_x += [x_start]
    vline_x = sorted(list(set(vline_x)))  #垂直线x坐标去重复
    #筛选单行文字   
    texts = msp.query('TEXT[layer=="%s"]' % layer)
    for text in texts:
        angle_text = round(text.dxf.rotation, 0)
        text_point = text.dxf.align_point
        if text_point is None:
            text_point = text.dxf.insert
        try:
#            print(text_point[0], text_point[1], text.dxf.text)
            if(x_min < text_point[0] < x_max and y_min < text_point[1] < y_max):
                if(angle_text % 180 == 0):
                    station += [(round(text_point[0], 0), text.dxf.text)]
                else:
                    a = (filter_strings(k_DK, text.dxf.text))[0]
                    b = (filter_strings(k_DK, text.dxf.text))[1]
                    po = float(a) + float(b) / 1000
#                    print(a, b, po)
                    position += [(text_point[0], po)]
        except Exception as e:
#            print(text_point[0])
            return float(text_point[0])
#            pass    
    station = list(dict.fromkeys(station))
    station = sorted(station, key = lambda x:x[0])
    position = list(dict.fromkeys(position))
    position = sorted(position, key = lambda x:x[0])
    for i in range(len(station)):
        for k in range(v_count, len(vline_x)):
            if(station[i][0] == round(vline_x[k], 0)):
#                print(station[i][1], position[i][1])
#                print(station[i][0])
                ws.cell(row = station_count + 2, column = 1, value = station[i][1])
#                print(station[i][1])
#                ws.cell(row = station_count + 2, column = 2, value = station[i][0])
#                v_count = k + 1
                station_count += 1
                break
    for m in range(len(position)):
        ws.cell(row = m + 2, column = 2, value = position[m][1])
#        print(position[m][1])
    
    save_path = Save
    wb.save(save_path)
    return 1


if __name__ == '__main__':
    aa = 'C:\\Users\\thinkpad\\Desktop\\线路条件转化\\test.dxf'
    bb = 'C:\\Users\\thinkpad\\Desktop\\线路条件转化\\设施.xlsx'
    cc = 0
    dd = 9661
    ee = -37
    ff = 80
    gg = "KbhZdmCzb"
    hh = "SK"
    run = station(aa, bb, cc, dd, ee, ff, gg, hh)