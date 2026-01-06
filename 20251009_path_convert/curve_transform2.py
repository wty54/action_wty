# -*- coding: utf-8 -*-
"""
Created on Mon Nov 18 13:52:48 2024

@author: thinkpad
"""

import ezdxf
import re
import openpyxl
from collections import Counter

#关键字查询与分割
def filter_strings(keywords, content):
    pattern = keywords
    result = []
    for string in content:
        if re.search(pattern, string):
#            string = list(string)
#            for i in range(len(string)):
#                if(string[i] == 'R' or string[i] == 'L'):
#                    if(string[i+1] == '-' and string[i+2] != ' '):
#                        result = string[i+2]
#                        for j in range(i+2, len(string), 1)
            string = re.sub(r'[{,},\\]', '', string)
            string = re.split(r'[-,\s,;]', string)     
           
            result.extend(string)
           
    return result

#关键字信息提取
def strings_position(string, keywords_L, keywords_R):
    for i in range(len(string)):
        if(string[i] == keywords_R or string[i] == keywords_L):
            break
    return i + 1

#重复信息筛选器
def layer_finder(ly):
    times = Counter(ly)
    most_common = times.most_common()[0]
    return most_common[0]

def ruler(Load, Save, x0, x1, y0, y1, t_flag, c_flag, k_L, k_R, Rx0, Rx1, Ry0, 
          Ry1, ruler_layer, STED_layer, st00):
#打开DWG文件
    dwg_file = Load
    doc = ezdxf.readfile(dwg_file)
    msp = doc.modelspace()
#曲线区域
    x_min = x0
    x_max = x1
    y_min = y0 
    y_max = y1
#公里标区域
    Rx_min = Rx0
    Rx_max = Rx1
    Ry_min = Ry0 
    Ry_max = Ry1
    ruler_asm = []  #公里标数组，用于排序
    STED_asm = []  #曲线起点终点数组
    ruler_mod = []  #公里标数组修正
    hline_count = 0  #曲线横线计数器（用于判断统计曲线方向）
    ruler_count = 0  #公里标计数器
    STED_mod = []  #曲线起点终点修正
    STED = []  #最终曲线起点终点
    STED_count = 0  #公里标输出excel计数器
    
    line_count = 0
    curve_count = 2 
    radius_count = 2
    distance_count = 2
    points = []
    ruler_points = []
    text_asm = []  #文字数组，用于排序
    color_asm = []  #颜色数组，用于筛选起终点20251009
#创建一个新的Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "曲线"
#创建表头
    ws['A1'] = "起点里程"
    ws['B1'] = "终点里程"
    ws['C1'] = "曲线方向"
    ws['D1'] = "曲线半径"
    ws['E1'] = "起缓和线长"
    ws['F1'] = "终缓和线长"
    ws['G1'] = "曲线全长"    

#曲线位置、方向统计(直线)
    if(c_flag == 0):
        lines = msp.query("LINE")
        for line in lines:
            x_start = round(line.dxf.start[0], 3)
            y_start = round(line.dxf.start[1], 3)
            x_end = round(line.dxf.end[0], 3)
            y_end = round(line.dxf.end[1], 3)
            if(x_min <= x_start <= x_max and y_min < y_start < y_max):   # 曲线位于长横线之内，不能触及长横线，否则需要删除长横线
                if(x_start > x_end):
                    points += [(x_end, y_end, x_start, y_start, line.dxf.layer)]
                else:
                    points += [(x_start, y_start, x_end, y_end, line.dxf.layer)]
#                line_layer.append(line.dxf.layer)
                line_count += 1
            elif(Rx_min <= x_start <= Rx_max and Ry_min < y_start <= Ry_max):
                ruler_points += [(x_start, y_start, x_end, y_end, line.dxf.layer)]
            
        points = list(dict.fromkeys(points))
        points = sorted(points, key = lambda x:x[0])
        ruler_points = list(dict.fromkeys(ruler_points))
        ruler_points = sorted(ruler_points, key = lambda x:x[0])
#        lineLayer = layer_finder(line_layer)
        #curve_line = points
        layer_line = [x for x in points if x[4] == STED_layer]
        ruler_vline = [x for x in ruler_points if x[4] == ruler_layer and x[0] == x[2]]
        h_line = [x for x in layer_line if x[1] == x[3]]
        h_line = list(dict.fromkeys(h_line))
        h_line = sorted(h_line, key = lambda x:x[0])
        h_line0 = []  #极值数组
        for row in h_line:
            h_line0.append(row[1]) 
        middle_value = round((max(h_line0) + min(h_line0)) / 2, 3)
        # 起终点垂线横坐标提取（用于构建起终点文字的横坐标）
        v_line = [x for x in layer_line if x[0] == x[2] and round(x[1], 3) == middle_value]
        v_line0 = []  #垂线横坐标数组
        for column in v_line:
            v_line0.append(column[0])
        v_line0 = list(dict.fromkeys(v_line0))
        v_line0 = sorted(v_line0)  # 用作起终点信息
        #        xst_0 = h_line[0][0]        
        # 计算并保存曲线方向信息20250920
        for cc in range(0, len(v_line0)-1, 2):
            for hcc in range(hline_count, len(h_line)):
                if(v_line0[cc]<=h_line[hcc][0]<=v_line0[cc+1] and v_line0[cc]<=h_line[hcc][2]<=v_line0[cc+1]):
                    if(h_line[hcc][1] - middle_value > 0):
                        ws.cell(row = curve_count, column = 3, value = 1)
                        curve_count += 1
                        hline_count = hcc
                        break
                    elif(h_line[hcc][1] - middle_value < 0):
                        ws.cell(row = curve_count, column = 3, value = 0)
                        curve_count += 1
                        hline_count = hcc
                        break
                    else:
                        continue
                else:
                    continue

#        curve_line = [x for x in layer_line if abs(round((x[3] - x[1]), 3)) == middle_value]
#        for i in range(len(curve_line)):
#            if(i % 2 == 0):
##                xst_asm += [curve_line[i][0]]  #比例尺法
##                ws.cell(row = int(i / 2) + 2, column = 1, value = curve_line[i][0])
#                if(curve_line[i][3] - curve_line[i][1] > 0):
#                    ws.cell(row = int(i / 2) + 2, column = 3, value = 1)
#                else:
#                    ws.cell(row = int(i / 2) + 2, column = 3, value = 0)

#曲线位置、方向统计(多段线)
    else:
        polylines = msp.query("LWPOLYLINE")
        for polyline in polylines:
            with polyline.points("xy") as points:
                points = sorted(points, key = lambda x:x[0])
                if(x_min < points[0][0] < x_max and y_min < points[0][1] < y_max):
                    for point_count in range(0, len(points), 4):
                        #曲线起点坐标计入表格
                        if(point_count + 4 > len(points)):
                            break
                        elif(points[point_count + 1][0] == points[point_count + 2][0]):
                            if((points[point_count + 2][1] + points[point_count + 1][1] - 2 * points[point_count][1]) > 0):
                                ws.cell(row = curve_count, column = 3, value = 1)
                            else:
                                ws.cell(row = curve_count, column = 3, value = 0)
                            curve_count += 1
                        else:
                            if(points[point_count + 2][1] - points[point_count + 1][1] > 0):
                                ws.cell(row = curve_count, column = 3, value = 1)
                            else:
                                ws.cell(row = curve_count, column = 3, value = 0)
                            curve_count += 1    
# 输入起终点信息    
    texts = msp.query("TEXT")
    for text in texts:
        if(Rx_min < text.dxf.insert[0] < Rx_max and Ry_min < text.dxf.insert[1] < Ry_max):
            ruler_asm += [(text.dxf.text, text.dxf.rotation, round(text.dxf.insert[0], 0), text.dxf.layer)]
            
        elif(x_min < text.dxf.insert[0] < x_max and y_min < text.dxf.insert[1] < y_max):
            STED_asm += [(text.dxf.text, text.dxf.rotation, round(text.dxf.insert[0], 0), text.dxf.layer, text.dxf.color)]  #20250930上海19号线增加文字纵坐标信息（未完成）
                
#建立公里标，输出公里标数组scale
    ruler_asm = list(dict.fromkeys(ruler_asm))
    ruler_asm = sorted(ruler_asm, key = lambda x:x[2])
    STED_asm = list(dict.fromkeys(STED_asm))
    STED_asm = sorted(STED_asm, key = lambda x:x[2])
    ruler_text = [(x[0], x[2]) for x in ruler_asm if x[3] == ruler_layer]  #公里标图层改形参
    for rr in ruler_text:
        ch1 = bool(re.search(r'[^0-9.]', rr[0]))
        if ch1:
            ruler_mod += [[st00, rr[1]]]
            st00 += 1
        else:
            ruler_mod += [[(st00 - 1) + int(rr[0]) / 10, rr[1]]]
#合成曲线起点终点
    STED_text = [(x[0].strip("+"), x[2], x[4]) for x in STED_asm if abs(round(x[1], 0)) == 90 and x[3] == STED_layer]  
    for color in STED_text:
        color_asm += [color[2]]
    color_common = layer_finder(color_asm)
    STED_text = [(x[0], x[1]) for x in STED_text if x[2] == color_common]
        #起点终点图层改形参
#筛选掉非起点终点的信息
    for ss in STED_text:
        ch2 = bool(re.search(r'[^0-9.]', ss[0]))
        if ch2:
            continue
        else:
            STED_mod += [[float(ss[0]), ss[1]]]
# 将起终点横坐标替换为垂线横坐标
    for vv in range(len(STED_mod)):
        STED_mod[vv][1] = v_line0[vv]
# 将公里标尺文字的横坐标替换为标尺线段横坐标
    for rr in range(len(ruler_mod)):
        ruler_mod[rr][1] = ruler_vline[rr][0]
# 按照公里标尺输出起终点
    for ST_count in range(len(STED_mod)):
        while(ruler_mod[ruler_count][1] <= STED_mod[ST_count][1]):
            ruler_count += 1
        STED += [float(ruler_mod[ruler_count - 1][0]) + STED_mod[ST_count][0] / 1000]
#        print(float(ruler_mod[ruler_count - 1][0]), STED_mod[ST_count][0])
#    print(STED)
#按文本输出曲线起点、终点
    for sd in range(len(STED)):
        if(sd % 2 == 0):
            ws.cell(row = STED_count + 2, column = 1, value = STED[sd])
        else:
            ws.cell(row = STED_count + 2, column = 2, value = STED[sd])
            STED_count += 1
    #曲线半径、全长统计（多行文字）
    if(t_flag == 0):
        mtexts = msp.query("MTEXT")
        for mtext in mtexts:
            if(mtext.get_rotation() == 0 and x_min < mtext.dxf.insert[0] < x_max and 
               y_min < mtext.dxf.insert[1] < y_max):
                text_asm += [(mtext.dxf.insert[0], mtext.text)] 
        text_asm = list(dict.fromkeys(text_asm))
        text_asm = sorted(text_asm, key = lambda x:x[0])
        for tt in range(len(text_asm)):
            text = [text_asm[tt][1]]
        #         print(text)
            radius = filter_strings((k_R + "-"), text)
            distance = filter_strings((k_L + "-"), text)
        #         print(radius)
            if(len(radius) == 0):
                if(len(distance) == 0):
                    continue
                else:
                    ws.cell(row = distance_count, column = 7,
                                 value = float(distance[strings_position(distance, k_L, k_R)]))
#                    length_asm += [float(distance[strings_position(distance, k_L, k_R)])]
                    distance_count += 1
            else:
                ws.cell(row = radius_count, column = 4, 
                             value = float(radius[strings_position(radius, k_L, k_R)]))
                radius_count += 1
    
    #曲线半径、全长统计（单行文字）
    else:
        texts = msp.query("TEXT")
        for text in texts:
            angle_text = text.dxf.rotation
            if(angle_text == 0 and x_min < text.dxf.insert[0] < x_max and 
               y_min < text.dxf.insert[1] < y_max):
                text_asm += [(text.dxf.insert[0], text.dxf.text)]
        text_asm = list(dict.fromkeys(text_asm))
        text_asm = sorted(text_asm, key = lambda x:x[0])
        for tt in range(len(text_asm)):
            text = [text_asm[tt][1]]
            radius = filter_strings((k_R + "-"), text)
            distance = filter_strings((k_L + "-"), text)
        #   print(radius)
            if(len(radius) == 0):
                if(len(distance) == 0):
                    continue
                else:
                    ws.cell(row = distance_count, column = 7, 
                                 value = float(distance[strings_position(distance, k_L, k_R)]))
#                    length_asm += [float(distance[strings_position(distance, k_L, k_R)])]
                    distance_count += 1
            else:
                ws.cell(row = radius_count, column = 4, 
                             value = float(radius[strings_position(radius, k_L, k_R)]))
                radius_count += 1

    save_path = Save
    wb.save(save_path)
    return 1
        
if __name__ == '__main__':
    aa = 'C:\\Users\\thinkpad\\Desktop\\线路条件转化\\test.dxf'
    bb = 'C:\\Users\\thinkpad\\Desktop\\线路条件转化\\右线曲线.xlsx'
    cc = 0
    dd = 9661
    ee = -300
    ff = -280
    gg = 0  # 单行文字——1  多行文字——0
    hh = 0  # 直线——0  多段线——1
    ii = 'L'
    jj = 'R'
    Rcc = 0
    Rdd = 9661
    Ree = -260
    Rff = -252
    LA_gg = "KbhLichYx"  # 公里标尺图层
    LA_hh = "KbhPmqxYx"  # 曲线图层
    st_ii = 1
#    jj = 'R'
#    kk = 19
    run = ruler(aa, bb, cc, dd, ee, ff, gg, hh, ii, jj, Rcc, Rdd, Ree, Rff, LA_gg,
                LA_hh, st_ii)