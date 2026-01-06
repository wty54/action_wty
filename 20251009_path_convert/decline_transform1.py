# -*- coding: utf-8 -*-
"""
Created on Mon Dec 18 10:45:39 2023

@author: thinkpad
"""

import ezdxf
import math
import openpyxl

def transform(Load, Save, x0, x1, y0, y1, t_flag, decst_0):
#打开DWG文件
#    dwg_file = 'C:\\Users\\thinkpad\\Desktop\\test3.dxf'
    dwg_file = Load
    doc = ezdxf.readfile(dwg_file)
    msp = doc.modelspace()
    lines = msp.query("LINE")
    

#遍历每个实体
#for entity in texts:
    #获取所有文字
#       rotation = entity.get_rotation()
#       position = entity.dxf.insert
#     print(entity.dxf.insert, entity.dxf.text)  
#    print(entity.text, '%d' % rotation, position)
#计数器设置
    podu_count = 2
    pochang_count = 2
    d_count = 0  #坡线计数器
    errorTimes = 0  #容错次数（缺省3次）
    vline_x = [] #垂直线列表
    d_line = [] #坡线列表
    d_text_list = [] #坡度文字列表
    l_text_list = [] #距离文字列表
#坡道数据区域
    #x_min = 0
    #x_max = 8627
    #y_min = -230 
    #y_max = -215
    x_min = x0
    x_max = x1
    y_min = y0 
    y_max = y1
#初始化
    sgn_decline = []
    route_start = decst_0  #起点里程
#计算坡度正负
    for line in lines:
        x0_line = round(line.dxf.start[0], 3)
        x1_line = round(line.dxf.end[0], 3)
        y0_line = round(line.dxf.start[1], 3)
        y1_line = round(line.dxf.end[1], 3)
        if(x_min <= x0_line <= x_max and x_min <= x1_line <= x_max and 
           y_min <= y0_line <= y_max and y_min <= y1_line <= y_max):
            if(x0_line < x1_line):
                d_line += [(x0_line, y0_line, x1_line, y1_line)]
            elif(x0_line == x1_line):
                vline_x += [x0_line]  #垂直线x坐标
            else:
                d_line += [(x1_line, y1_line, x0_line, y0_line)]
        
#坡度线排序，垂直线去重
    vline_x = sorted(list(set(vline_x)))
    d_line = list(dict.fromkeys(d_line)) #去除重复的
    d_line = sorted(d_line, key = lambda x:x[0]) #从小到大排序
#    dline_y0max = round(max(d_line, key = lambda x:x[1])[1], 4)
#    dline_y0min = round(min(d_line, key = lambda x:x[1])[1], 4)
#    dline_y1max = max(d_line, key = lambda x:x[3])[3]
#    dline_y1min = min(d_line, key = lambda x:x[3])[3]
    
#    if(len(vline_x) != len(d_line) + 1):
#        print("垂直线或坡道线数量有误，请检查坡道")
#        return -1
#双垂直线筛选（不适用于无起始线）    
#    for i in range(len(vline_x) - 1):
#        for j in range(d_count ,len(d_line)):
#            if(round(d_line[d_count][1], 4) == round(d_line[d_count][3], 4) 
#            == dline_y0max or round(d_line[d_count][1], 4) == round(d_line[d_count][3], 4) 
#            == dline_y0min):
#                d_count += 1
#                continue
#            while errorTimes < 3:
#                if(d_line[j][0] == vline_x[i] and d_line[j+errorTimes][2] == vline_x[i+1]):
#                    line_decline = (d_line[j][3] - d_line[j][1]) / (d_line[j][2] - d_line[j][0])
#                    if(line_decline == 0):
#                        sgn_decline += [0]
#                    else:
#                        sgn_decline += [int(math.copysign(1, line_decline))]
#                    d_count += 1
#                    errorTimes = 0
#                    break
#                errorTimes += 1
#                d_count += 1
#            if(errorTimes >= 3):
#                print("坡道位置'%f'处出现错误"  % d_line[j][0])
#                return -1
#            break

#垂直线终点筛选
    for i in range(len(vline_x)):
        if(d_count > len(d_line)):
            break
        while d_count < len(d_line):
#            if(round(d_line[d_count][1], 4) == round(d_line[d_count][3], 4) 
#            == dline_y0max or round(d_line[d_count][1], 4) == round(d_line[d_count][3], 4) 
#            == dline_y0min):
#                d_count += 1
#                continue
            if(d_line[d_count+errorTimes][2] < vline_x[i]):
                d_count += 1
                errorTimes = 0
            elif(d_line[d_count+errorTimes][2] == vline_x[i]):
                line_decline = (d_line[d_count+errorTimes][3] - d_line[d_count+errorTimes][1]) / (d_line[d_count+errorTimes][2] - d_line[d_count+errorTimes][0])
                if(line_decline == 0):
                    sgn_decline += [0]
                else:
                    sgn_decline += [int(math.copysign(1, line_decline))]
                d_count = d_count + errorTimes + 1
                errorTimes = 0
                break
            else:
                errorTimes += 1
#                j = d_count - 1
                if(errorTimes >= 3):
#                    print("坡道位置'%f'处出现错误"  % d_line[j][0])
                    errorTimes = 0
                    break


#创建一个新的Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "坡度"

    #创建表头
    ws['A1'] = "起点里程"
    ws['B1'] = "终点里程"
    #ws['C1'] = "X2"
    #ws['D1'] = "Y2"
    ws['C1'] = "坡度"
    ws['D1'] = "坡长"
    ws['E1'] = "长短链里程"
    ws['F1'] = "长短链"
    #属性为多行文字
    if(t_flag == 1):
        mtexts = msp.query("MTEXT")
        for mtext in mtexts:
            if(round(mtext.get_rotation(), 0) % 180 == 0 and x_min < mtext.dxf.insert[0] < x_max):
                if((y_min + y_max) / 2 < mtext.dxf.insert[1] < y_max):
                    ws.cell(row = podu_count, column = 3, value = float(mtext) * sgn_decline[podu_count-2])
                    podu_count += 1
                elif(y_min < mtext.dxf.insert[1] < (y_min + y_max) / 2):
                    ws.cell(row = pochang_count, column = 1, value = route_start)
                    ws.cell(row = pochang_count, column = 4, value = float(mtext))
                    ws.cell(row = pochang_count, column = 2, value = 
                            float(mtext) / 1000 + route_start)
                    pochang_count += 1
                    route_start += float(mtext) / 1000
    
    #属性为单行文字
    else:
        texts = msp.query("TEXT")
        for text in texts:
            angle_text = round(text.dxf.rotation, 0)
            x_text = text.dxf.insert[0]
            y_text = text.dxf.insert[1]
            if(angle_text % 180 == 0 and x_min < x_text < x_max):
                if((y_min + y_max) / 2 < y_text < y_max):
                    d_text_list += [(text.dxf.text, x_text, y_text)]
                elif(y_min < y_text < (y_min + y_max) / 2):
                    l_text_list += [(text.dxf.text, x_text, y_text)]
                else:
                    continue
        d_text_list = list(dict.fromkeys(d_text_list)) #去除重复的
        d_text_list = sorted(d_text_list, key = lambda x:x[1]) #从小到大排序
        l_text_list = list(dict.fromkeys(l_text_list)) #去除重复的
        l_text_list = sorted(l_text_list, key = lambda x:x[1]) #从小到大排序
        for d_text in d_text_list:
            ws.cell(row = podu_count, column = 3, value = abs(float(d_text[0])) * sgn_decline[podu_count-2])
#                    print(float(text.dxf.text), text.dxf.insert[0]) #调试用
            podu_count += 1
        for l_text in l_text_list:
            ws.cell(row = pochang_count, column = 1, value = route_start)
            ws.cell(row = pochang_count, column = 4, value = float(l_text[0]))
            ws.cell(row = pochang_count, column = 2, value = float(l_text[0]) / 1000 + route_start)
#                    print(float(text.dxf.text), text.dxf.insert[0]) #调试用
            pochang_count += 1
            route_start += float(l_text[0]) / 1000
                
    #保存EXCEL文件
    #    wb.save('C:\\Users\\thinkpad\\Desktop\\test3.xlsx')
    wb.save(Save)
    return 1  
    

if __name__ == '__main__':
    aa = 'C:\\Users\\thinkpad\\Desktop\\线路条件转化\\test左线.dxf'
    bb = 'C:\\Users\\thinkpad\\Desktop\\线路条件转化\\坡道左线.xlsx'
    cc = -162
    dd = 7189
    ee = -216
    ff = -204
    gg = 0
    hh = 2.2
    run = transform(aa, bb, cc, dd, ee, ff, gg, hh)    
#   